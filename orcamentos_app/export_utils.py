"""
Geração da planilha .xlsx final: mapa comparativo, aba de revisão e aba de fontes.
"""
import io
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _row_key(row):
    if row.get("_key") is not None:
        return row.get("_key")
    if row.get("numero_item"):
        return ("num", str(row["numero_item"]).strip())
    return ("desc", (row.get("descricao") or "").strip().lower())


def _thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def build_excel(rows, matrix, empresas, review, sources):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapa Comparativo"

    n_cols = 4 + len(empresas)  # Nº Item + Descrição + UF + QTD + empresas

    # ── Linha 1: título "MAPA COMPARATIVO DE PREÇOS" mesclado e em destaque ──
    titulo_fill = PatternFill(start_color="0D2137", end_color="0D2137", fill_type="solid")
    titulo_font = Font(color="FFFFFF", bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    titulo_cell = ws.cell(row=1, column=1, value="MAPA COMPARATIVO DE PREÇOS")
    titulo_cell.fill = titulo_fill
    titulo_cell.font = titulo_font
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Linha 2: cabeçalho das colunas ──
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    min_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    low_conf_font = Font(italic=True, color="9C5700")
    borda = _thin_border()
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    headers = ["Nº Item", "Descrição", "UF", "QTD"] + empresas
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_wrap
        cell.border = borda
    ws.row_dimensions[2].height = 36

    for r_idx, row in enumerate(rows, start=3):
        # ── Nº Item (col 1) ──
        c = ws.cell(row=r_idx, column=1, value=row.get("numero_item"))
        c.alignment = center_wrap
        c.border = borda

        # ── Descrição (col 2) ──
        c = ws.cell(row=r_idx, column=2, value=row.get("descricao"))
        c.alignment = left_wrap
        c.border = borda

        key = _row_key(row)
        cell_values = matrix.get(key, {})
        unidade = row.get("unidade")
        quantidade = row.get("quantidade")
        if unidade is None or quantidade is None:
            for empresa in empresas:
                info = cell_values.get(empresa) or {}
                if unidade is None and info.get("unidade") is not None:
                    unidade = info.get("unidade")
                if quantidade is None and info.get("quantidade") is not None:
                    quantidade = info.get("quantidade")
                if unidade is not None and quantidade is not None:
                    break

        # ── UF (col 3) ──
        c = ws.cell(row=r_idx, column=3, value=unidade)
        c.alignment = center_wrap
        c.border = borda

        # ── QTD (col 4) ──
        c = ws.cell(row=r_idx, column=4, value=quantidade)
        c.alignment = center_wrap
        c.border = borda

        precos_validos = []
        for c_idx, empresa in enumerate(empresas, start=5):
            info = cell_values.get(empresa)
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = center_wrap
            cell.border = borda
            if info and info.get("preco_unitario") is not None:
                cell.value = info["preco_unitario"]
                cell.number_format = 'R$ #,##0.00'
                if info.get("confianca") == "média":
                    cell.font = low_conf_font
                precos_validos.append((c_idx, info["preco_unitario"]))

        if precos_validos:
            min_col, _ = min(precos_validos, key=lambda x: x[1])
            ws.cell(row=r_idx, column=min_col).fill = min_fill

    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 50
    ws.column_dimensions[get_column_letter(3)].width = 10
    ws.column_dimensions[get_column_letter(4)].width = 12
    for col in range(5, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.freeze_panes = "E3"

    # Aba de revisão manual (casamentos incertos por descrição ou por inconsistência de número)
    ws_rev = wb.create_sheet("Revisar Casamentos")
    ws_rev.append(["Tipo", "Nº Item", "Descrição Nova", "Casou Com", "Score de Similaridade"])
    for cell in ws_rev[1]:
        cell.font = header_font
        cell.fill = header_fill
    for r in review:
        ws_rev.append([
            r.get("tipo", ""),
            r.get("numero_item"),
            r["descricao_nova"],
            r["casou_com"],
            r["score"],
        ])
    ws_rev.column_dimensions["A"].width = 30
    ws_rev.column_dimensions["B"].width = 12
    ws_rev.column_dimensions["C"].width = 45
    ws_rev.column_dimensions["D"].width = 45
    ws_rev.column_dimensions["E"].width = 20

    # Aba de rastreabilidade (empresa -> arquivo -> fonte -> localizacao)
    ws_src = wb.create_sheet("Fontes")
    ws_src.append(["Empresa", "Arquivo de Origem", "Fonte de Extração", "Localização", "Nº Item", "Descrição"])
    for cell in ws_src[1]:
        cell.font = header_font
        cell.fill = header_fill
    for src in sources:
        if isinstance(src, dict):
            ws_src.append([
                src.get("empresa"),
                src.get("arquivo"),
                src.get("fonte_extracao"),
                src.get("origem"),
                src.get("numero_item"),
                src.get("descricao"),
            ])
        else:
            empresa, arquivo = src
            ws_src.append([empresa, arquivo, None, None, None, None])
    ws_src.column_dimensions["A"].width = 35
    ws_src.column_dimensions["B"].width = 45
    ws_src.column_dimensions["C"].width = 22
    ws_src.column_dimensions["D"].width = 22
    ws_src.column_dimensions["E"].width = 12
    ws_src.column_dimensions["F"].width = 50

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
