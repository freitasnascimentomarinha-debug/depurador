"""Checklist executavel de saude para deploy Streamlit."""
from __future__ import annotations

import importlib
import subprocess
import sys

REQUIRED_MODULES = [
    "streamlit",
    "pandas",
    "pytesseract",
    "PIL",
    "docx",
    "openpyxl",
    "rapidfuzz",
    "requests",
    "pdfplumber",
]

OPTIONAL_MODULES = [
    "fitz",
]

PROJECT_MODULES = [
    "normalize_utils",
    "structured_extract",
    "confidence",
    "extract_utils",
    "match_utils",
    "export_utils",
    "db_utils",
    "app",
]


def _check_module(name: str) -> tuple[bool, str]:
    try:
        importlib.import_module(name)
        return True, "ok"
    except Exception as exc:  # pragma: no cover - script de diagnostico
        return False, f"{type(exc).__name__}: {exc}"


def _check_tesseract() -> tuple[bool, str]:
    try:
        proc = subprocess.run(
            ["tesseract", "--version"],
            check=True,
            capture_output=True,
            text=True,
        )
        versao = (proc.stdout or proc.stderr or "").splitlines()[0].strip()
        return True, versao or "ok"
    except Exception as exc:  # pragma: no cover - script de diagnostico
        return False, f"{type(exc).__name__}: {exc}"


def _check_tesseract_langs(required: list[str]) -> tuple[bool, str]:
    try:
        proc = subprocess.run(
            ["tesseract", "--list-langs"],
            check=True,
            capture_output=True,
            text=True,
        )
        saida = (proc.stdout or "")
        langs = {
            l.strip()
            for l in saida.splitlines()
            if l.strip() and not l.lower().startswith("list of available languages")
        }
        faltando = [l for l in required if l not in langs]
        if faltando:
            return False, f"Idiomas ausentes: {', '.join(faltando)}"
        return True, f"Idiomas OK: {', '.join(required)}"
    except Exception as exc:  # pragma: no cover - script de diagnostico
        return False, f"{type(exc).__name__}: {exc}"


def _check_extraction_engines() -> list[tuple[str, bool, str]]:
    checks = []

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["item", "descricao", "qtd", "valor"])
        ws.append([1, "Item teste", 2, 10.5])
        checks.append(("openpyxl tabela", True, "ok"))
    except Exception as exc:
        checks.append(("openpyxl tabela", False, f"{type(exc).__name__}: {exc}"))

    try:
        from docx import Document
        doc = Document()
        table = doc.add_table(rows=2, cols=2)
        table.cell(0, 0).text = "item"
        table.cell(0, 1).text = "descricao"
        table.cell(1, 0).text = "1"
        table.cell(1, 1).text = "Item teste"
        checks.append(("python-docx tabela", True, "ok"))
    except Exception as exc:
        checks.append(("python-docx tabela", False, f"{type(exc).__name__}: {exc}"))

    try:
        import pdfplumber
        _ = pdfplumber.open
        checks.append(("pdfplumber leitura", True, "ok"))
    except Exception as exc:
        checks.append(("pdfplumber leitura", False, f"{type(exc).__name__}: {exc}"))

    return checks


def main() -> int:
    print(f"Python: {sys.version}")
    print("--- Dependencias ---")
    has_error = False
    for module in REQUIRED_MODULES:
        ok, msg = _check_module(module)
        status = "OK" if ok else "ERRO"
        print(f"[{status}] {module}: {msg}")
        has_error = has_error or (not ok)

    print("--- Dependencias opcionais ---")
    for module in OPTIONAL_MODULES:
        ok, msg = _check_module(module)
        status = "OK" if ok else "AVISO"
        print(f"[{status}] {module}: {msg}")

    print("--- OCR (Tesseract) ---")
    ok_tess, msg_tess = _check_tesseract()
    print(f"[{'OK' if ok_tess else 'ERRO'}] tesseract: {msg_tess}")
    has_error = has_error or (not ok_tess)

    ok_langs, msg_langs = _check_tesseract_langs(["por", "eng"])
    print(f"[{'OK' if ok_langs else 'ERRO'}] idiomas OCR: {msg_langs}")
    has_error = has_error or (not ok_langs)

    print("--- Motores de extracao (regras antes da IA) ---")
    for nome, ok, msg in _check_extraction_engines():
        status = "OK" if ok else "ERRO"
        print(f"[{status}] {nome}: {msg}")
        has_error = has_error or (not ok)

    print("--- Modulos do projeto ---")
    for module in PROJECT_MODULES:
        ok, msg = _check_module(module)
        status = "OK" if ok else "ERRO"
        print(f"[{status}] {module}: {msg}")
        has_error = has_error or (not ok)

    if has_error:
        print("\nResultado: FALHA. Corrija os erros acima antes do deploy.")
        return 1

    print("\nResultado: OK. Ambiente pronto para deploy.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
